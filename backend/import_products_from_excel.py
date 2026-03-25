from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.database_dynamodb import create_category, create_product, get_all_categories, get_all_products

WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "Untitled spreadsheet.xlsx"


def normalize_name(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    return normalize_name(value).casefold()


def number_or_default(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except Exception:
        return default


def build_description(item_no: str) -> str:
    item_no = normalize_name(item_no)
    return f"Item no : {item_no}" if item_no else ""


def resolve_price_and_mrp(primary: Any, secondary: Any) -> tuple[float, float]:
    values = [number_or_default(primary, 0.0), number_or_default(secondary, 0.0)]
    values = [v for v in values if v > 0]
    if not values:
        return 0.0, 0.0
    if len(values) == 1:
        return values[0], values[0]
    return min(values), max(values)


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [normalize_name(h) for h in rows[0]]

    categories = get_all_categories()
    category_lookup = {(normalize_key(cat.get("name")), cat.get("parent_id")): cat for cat in categories}
    products = get_all_products()
    existing_product_ids = {normalize_key(p.get("product_id")) for p in products if p.get("product_id")}

    created_categories = 0
    created_products = 0
    skipped_products = 0

    for row in rows[1:]:
        if not row or not normalize_name(row[1]):
            continue

        product_code = normalize_name(row[0])
        if product_code and normalize_key(product_code) in existing_product_ids:
            skipped_products += 1
            continue

        category_name = normalize_name(row[3])
        subcategory_name = normalize_name(row[4])

        if not category_name:
            skipped_products += 1
            continue

        category_key = (normalize_key(category_name), None)
        category = category_lookup.get(category_key)
        if not category:
            category_id = create_category({
                "name": category_name,
                "icon": None,
                "image_url": None,
                "parent_id": None,
                "display_order": 0,
            })
            category = {
                "id": category_id,
                "name": category_name,
                "parent_id": None,
            }
            category_lookup[category_key] = category
            created_categories += 1

        final_category_id = int(category["id"])
        subcategory_id = None
        if subcategory_name:
            subcategory_key = (normalize_key(subcategory_name), int(category["id"]))
            subcategory = category_lookup.get(subcategory_key)
            if not subcategory:
                created_id = create_category({
                    "name": subcategory_name,
                    "icon": None,
                    "image_url": None,
                    "parent_id": int(category["id"]),
                    "display_order": 0,
                })
                subcategory = {
                    "id": created_id,
                    "name": subcategory_name,
                    "parent_id": int(category["id"]),
                }
                category_lookup[subcategory_key] = subcategory
                created_categories += 1
            subcategory_id = int(subcategory["id"])
            final_category_id = subcategory_id

        price, mrp = resolve_price_and_mrp(row[8], row[9])
        product_payload = {
            "product_id": product_code or None,
            "name": normalize_name(row[1]),
            "unit": normalize_name(row[2]) or "pcs",
            "category_id": final_category_id,
            "subcategory_id": subcategory_id,
            "brand": normalize_name(row[5]) or None,
            "cost_price": number_or_default(row[6], 0.0),
            "mrp": mrp,
            "price": price,
            "stock": int(number_or_default(row[10], 0.0)),
            "description": build_description(normalize_name(row[11])),
            "image": "",
            "discount": 0.0,
            "out_of_stock": int(number_or_default(row[10], 0.0)) <= 0,
        }

        created_id = create_product(product_payload)
        if created_id:
            created_products += 1
            if product_code:
                existing_product_ids.add(normalize_key(product_code))
        else:
            skipped_products += 1

    print({
        "rows": max(0, len(rows) - 1),
        "created_categories": created_categories,
        "created_products": created_products,
        "skipped_products": skipped_products,
    })


if __name__ == "__main__":
    main()
