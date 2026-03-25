from __future__ import annotations

from pathlib import Path
from typing import Any

import boto3
from openpyxl import load_workbook

from backend.database_dynamodb import (
    REGION,
    PRODUCTS_TABLE,
    CATEGORIES_TABLE,
    _resource_kwargs,
    _scan_all,
    create_category,
    create_product,
    get_all_categories,
)

WORKBOOK_PATH = Path(__file__).resolve().parents[1] / "ak_store.xlsx"
STACK_BUCKET = "ak-store-api-prod-007222077181-uploads"


def normalize(value: Any, default: str = "") -> str:
    text = str(value or "").strip()
    return text or default


def number(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except Exception:
        return default


def clear_table(table_name: str) -> int:
    dynamodb = boto3.resource(**_resource_kwargs("dynamodb"))
    table = dynamodb.Table(table_name)
    items = _scan_all(table_name)
    deleted = 0
    with table.batch_writer() as batch:
        for item in items:
            batch.delete_item(Key={"id": item["id"]})
            deleted += 1
    return deleted


def clear_bucket(bucket_name: str) -> int:
    s3 = boto3.resource(**_resource_kwargs("s3"))
    bucket = s3.Bucket(bucket_name)
    keys = [{"Key": obj.key} for obj in bucket.objects.all()]
    if not keys:
        return 0
    deleted = 0
    for i in range(0, len(keys), 1000):
        chunk = keys[i:i + 1000]
        bucket.delete_objects(Delete={"Objects": chunk, "Quiet": True})
        deleted += len(chunk)
    return deleted


def ensure_category(name: str, parent_id: int | None, cache: dict[tuple[str, int | None], int]) -> int:
    key = (name.casefold(), parent_id)
    if key in cache:
        return cache[key]
    category_id = create_category({
        "name": name,
        "icon": None,
        "image_url": None,
        "parent_id": parent_id,
        "display_order": 0,
    })
    cache[key] = int(category_id)
    return int(category_id)


def import_sheet() -> dict[str, int]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    category_cache: dict[tuple[str, int | None], int] = {}
    for category in get_all_categories():
        category_cache[(normalize(category.get("name")).casefold(), category.get("parent_id"))] = int(category["id"])

    imported = 0
    skipped = 0
    for row in rows[1:]:
        product_code = normalize(row[1], "NA")
        product_name = normalize(row[2])
        if not product_name:
            skipped += 1
            continue

        main_category_name = normalize(row[4], "NA")
        sub_category_name = normalize(row[5])
        manufacturer = normalize(row[6], "NA")
        sales_price = number(row[7], 0.0)
        purchase_price = number(row[8], 0.0)
        mrp = number(row[9], sales_price)
        item_code = normalize(row[10], "NA")

        main_category_id = ensure_category(main_category_name, None, category_cache)
        subcategory_id = None
        final_category_id = main_category_id
        if sub_category_name:
            subcategory_id = ensure_category(sub_category_name, main_category_id, category_cache)
            final_category_id = subcategory_id

        payload = {
            "product_id": product_code,
            "name": product_name,
            "price": sales_price,
            "cost_price": purchase_price,
            "mrp": mrp if mrp > 0 else sales_price,
            "discount": 0.0,
            "description": f"Item code: {item_code}",
            "unit": "NA",
            "image": "",
            "category_id": final_category_id,
            "subcategory_id": subcategory_id,
            "stock": 0,
            "out_of_stock": False,
            "brand": manufacturer,
            "catch": "NA",
            "product_type": "NA",
            "mfg_date": "NA",
            "country_of_origin": "NA",
            "manufacturer_name": manufacturer,
            "manufacturer_address": "NA",
            "seller_name": "NA",
            "customer_care_details": "NA",
            "disclaimer": "NA",
        }
        if create_product(payload):
            imported += 1
        else:
            skipped += 1

    return {"imported": imported, "skipped": skipped}


def main() -> None:
    deleted_products = clear_table(PRODUCTS_TABLE)
    deleted_categories = clear_table(CATEGORIES_TABLE)
    deleted_s3_objects = clear_bucket(STACK_BUCKET)
    result = import_sheet()
    print({
        "deleted_products": deleted_products,
        "deleted_categories": deleted_categories,
        "deleted_s3_objects": deleted_s3_objects,
        **result,
    })


if __name__ == "__main__":
    main()
